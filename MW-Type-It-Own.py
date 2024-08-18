from PyQt6.QtWidgets import QApplication, QWidget, QGridLayout, QPushButton, QVBoxLayout, \
                            QLabel, QMessageBox, QFrame, QSpacerItem, QSizePolicy
from PyQt6.QtCore import QTimer, QEvent, QCoreApplication, Qt
from PyQt6.QtGui import QIcon, QFont, QFontDatabase
from openpyxl import load_workbook
import pyautogui
from time import sleep
import os, sys
import keyboard
from refresh_button import RefreshButton

# Suppress Qt warnings by setting environment variable
os.environ['QT_LOGGING_RULES'] = "qt.qpa.window=false"

def get_resource_path(relative_path):
    """ Get absolute path to resource, works for development and for PyInstaller """
    if hasattr(sys, '_MEIPASS'):
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        return os.path.join(sys._MEIPASS, relative_path)
    else:
        # PyCharm or other IDEs
        return os.path.join(os.path.dirname(__file__), relative_path)

class KeyPressEvent(QEvent):
    def __init__(self, key):
        super().__init__(QEvent.Type.User)
        self.key = key

class MyWidget(QWidget):
    def __init__(self):
        super().__init__()

        # Check if the required Excel file exists
        self.excel_file_path = "button_names.xlsx"
        if not os.path.exists(self.excel_file_path):
            self.show_error_message("Error", "The required Excel file is missing !!!")
            sys.exit()

        self.setWindowTitle("Type-It-Own")
        favicon_path = get_resource_path("static/wf_favicon_48x48.webp")
        self.setWindowIcon(QIcon(favicon_path))

        # Create the main layout for the widget
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(0, 10, 0, 10)  # Set margins for the entire layout

        font_com_path = get_resource_path('static/Baskervville-Regular.ttf')
        font_com_id = QFontDatabase.addApplicationFont(font_com_path)
        if font_com_id != -1:
            font_com_name = QFontDatabase.applicationFontFamilies(font_com_id)[0]
        # Set up the top label with 100px height
        self.top_label = QLabel('WELLS FARGO', self)
        self.top_label.setFixedHeight(75)
        self.top_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        if font_com_id != -1:
            self.top_label.setFont(QFont(font_com_name))
        self.top_label.setStyleSheet("font-weight:bold; font-size:30px;"
                                     "background-color: #d71e28; color:white;"
                                     "letter-spacing: 1.5px; border: 8px solid #ffcd41")
        main_layout.addWidget(self.top_label)

        # Add a spacer of 20px between the top label and the new label
        spacer_between_labels = QSpacerItem(20, 10, QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Fixed)
        main_layout.addSpacerItem(spacer_between_labels)

        font_app_path = get_resource_path('static/SpecialElite-Regular.ttf')
        font_app_id = QFontDatabase.addApplicationFont(font_app_path)
        if font_app_id != -1:
            font_app_name = QFontDatabase.applicationFontFamilies(font_app_id)[0]

        # Set up the new label below the top label
        self.middle_label = QLabel('Type-It-Own', self)
        self.middle_label.setFixedHeight(50)
        if font_app_id != -1:
            self.middle_label.setFont(QFont(font_app_name))
        self.middle_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.middle_label.setStyleSheet("font-weight:bold; font-size:30px;"
                                        "border-top: 3px solid #585759;"
                                        "border-bottom: 3px solid #585759;"
                                        "border-left: 50px solid #585759;"
                                        "border-right: 50px solid #585759;")
        #0727f2;
        main_layout.addWidget(self.middle_label)

        # Add a spacer to position the buttons correctly
        spacer_top = QSpacerItem(20, 30, QSizePolicy.Policy.Maximum, QSizePolicy.Policy.Expanding)
        main_layout.addSpacerItem(spacer_top)

        
        # Create an instance of RotatingButton with text "Refresh Data"
        self.rotating_button = RefreshButton(self)
        self.rotating_button.refresh_excel_data = self.load_and_update_data
        # Add the rotating button to the layout
        # self.apply_button_style(self.rotating_button, bg_color="grey", hr_color="#3FA2F6")
        main_layout.addWidget(self.rotating_button)

        # Create a QLabel to display the text from the Excel file
        self.top_label = QLabel("   Select a button", self)
        self.top_label.setStyleSheet("font-weight:bold; font-size:15px; padding-bottom: 10px;")
        main_layout.addWidget(self.top_label)

        self.grid_layout = QGridLayout()
        self.grid_layout.setContentsMargins(20, 0, 20, 0)  # Add 20 pixels of padding to the left and right
        self.grid_layout.setSpacing(10)





        main_layout.addLayout(self.grid_layout)
                              
        # Add another spacer below the buttons to maintain the position
        spacer_bottom = QSpacerItem(15, 15)
        main_layout.addSpacerItem(spacer_bottom)

        # Set up the bottom label with 50px height
        self.bottom_label = QLabel('#MW ®', self)
        self.bottom_label.setFixedHeight(30)
        self.bottom_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.bottom_label.setStyleSheet("background-color: lightgrey; font-size: 18px; font-weight: bold;")
        main_layout.addWidget(self.bottom_label)

        #####
        self.load_and_update_data()
        # Setup global hotkeys
        self.setup_hotkeys()


    def draw_vertical_line(self):
        # Create a vertical line (QFrame)
        line = QFrame(self)
        line.setFrameShape(QFrame.Shape.VLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        line.setLineWidth(3)  # Set the line width (thickness)
        line.setStyleSheet("border: 2px solid grey;")  # Set the color and thickness
        return line
        # grid_layout.addWidget(line, 0, col_num, 13, 1)

    def apply_button_style(self, button, bg_color, hr_color):
        """Apply a consistent style to the given button."""
        button.setStyleSheet(f"""
            QPushButton {{
                color: black;           /* Text color */
                background-color: {bg_color} ; /* Background color */
                font-size: 13px;        /* Font size */
                font-weight: bold;      /* Font weight */
                font-family: Arial;     /* Font family */
                border-radius: 10px;    /* Rounded corners */
                padding: 10px 20px;     /* Padding */
                text-align: left;
            }}
            QPushButton:hover {{
                background-color: {hr_color}; /* Background color on hover */
            }}
            QPushButton:pressed {{
                background-color: navy; /* Background color when pressed */
                color: white;
            }}
        """)

    def load_and_update_data(self):
        self.clear_buttons()
        self.button_data = self.read_button_names_from_excel(self.excel_file_path)
        # print("Excel data reloaded.")
        # print(self.button_data[-1])
        self.create_buttons_from_data()

    def create_buttons_from_data(self):
        # For 0 to 9 buttons
        for i in range(10):
            index = 0 + i
            button = QPushButton(self)
            self.apply_button_style(button, bg_color="#96C9F4", hr_color="#3FA2F6")
            if index < len(self.button_data):
                if self.button_data[index]['name'] is not None and self.button_data[index]['name'].find("N/A") == -1:
                    button_text = self.button_data[index]['name']
                    button_row = self.button_data[index]['row']
                    button.setText(button_text)
                    button.clicked.connect(lambda checked, row=button_row: self.on_button_click(row))
                else:
                    button.setText(self.button_data[index]['name'])
                    button.setEnabled(False)
                    self.apply_button_style(button, bg_color="grey", hr_color="#3FA2F6")
            else:
                button.setText("N/A")
                button.setEnabled(False)
                self.apply_button_style(button, bg_color="grey", hr_color="#3FA2F6")
            self.grid_layout.addWidget(button, i, 0)

        self.grid_layout.addWidget(self.draw_vertical_line(), 0, 1, 13, 1)

        # For A to M buttons
        for i in range(13):
            index = 10 + (i + 0 * 13)
            button = QPushButton(self)
            self.apply_button_style(button, bg_color="#82dbb8", hr_color="#53ab8b")
            if index < len(self.button_data):
                if self.button_data[index]['name'] is not None and self.button_data[index]['name'].find("N/A") == -1:
                    button_text = self.button_data[index]['name']
                    button_row = self.button_data[index]['row']
                    button.setText(button_text)
                    button.clicked.connect(lambda checked, row=button_row: self.on_button_click(row))
                else:
                    button.setText(self.button_data[index]['name'])
                    button.setEnabled(False)
                    self.apply_button_style(button, bg_color="grey", hr_color="#3FA2F6")
            else:
                button.setText("N/A")
                button.setEnabled(False)
                self.apply_button_style(button, bg_color="grey", hr_color="#3FA2F6")
            self.grid_layout.addWidget(button, i, 2)

        self.grid_layout.addWidget(self.draw_vertical_line(), 0, 3, 13, 1)

        # For N to Z buttons
        for i in range(13):
            index = 10 + (i + 1 * 13)
            button = QPushButton(self)
            self.apply_button_style(button, bg_color="#82dbb8", hr_color="#53ab8b")
            if index < len(self.button_data):
                if self.button_data[index]['name'] is not None and self.button_data[index]['name'].find("N/A") == -1:
                    button_text = self.button_data[index]['name']
                    button_row = self.button_data[index]['row']
                    button.setText(button_text)
                    button.clicked.connect(lambda checked, row=button_row: self.on_button_click(row))
                else:
                    button.setText(self.button_data[index]['name'])
                    button.setEnabled(False)
                    self.apply_button_style(button, bg_color="grey", hr_color="#3FA2F6")
            else:
                button.setText("N/A")
                button.setEnabled(False)
                self.apply_button_style(button, bg_color="grey", hr_color="#3FA2F6")
            self.grid_layout.addWidget(button, i, 4)

    def clear_buttons(self):
        # Remove all widgets (buttons) from the grid layout
        for i in reversed(range(self.grid_layout.count())):
            widget = self.grid_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()

    def read_button_names_from_excel(self, file_path):
        workbook = load_workbook(file_path, data_only=True)
        sheet = workbook.active
        button_data = []

        for row in range(1, 37):  # Assuming up to 36 rows
            col1 = sheet.cell(row=row, column=1).value
            col2 = sheet.cell(row=row, column=2).value
            col3 = sheet.cell(row=row, column=3).value
            if col1 is not None and col2 is not None:
                button_name = f"({col1})   {col2} "
                button_data.append({'name': button_name, 'row': row, 'col3_text': col3, 'label': col2})
            elif col1 is not None and col2 is None:
                button_name = f"({col1})   N/A"
                button_data.append({'name': button_name, 'row': row, 'col3_text': col3, 'label': col2})
        return button_data

    def on_button_click(self, row):
        # Retrieve the text from the third column of the clicked button's row
        for data in self.button_data:
            if data['row'] == row:
                text_to_display = data['col3_text']
                self.top_label.setText(f"  Selected: \"{data['label']}\"")
                # Start typing text after 2 seconds delay
                QTimer.singleShot(2000, lambda: self.simulate_typing(text_to_display))
                break

    def simulate_typing(self, text):
        sleep(1)
        pyautogui.write(text)

    def setup_hotkeys(self):
        self.hotkey_ids = []
        # Set up global hotkeys for each letter (A-Z)
        for i in range(65, 91):  # ASCII for A-Z
            letter = chr(i)
            hotkey = f'ctrl+shift+{letter.lower()}'
            keyboard.add_hotkey(hotkey, lambda l=letter: self.post_key_press_event(l))
            self.hotkey_ids.append(hotkey)
        
        # Set up global hotkeys for each number (0-9)
        for i in range(10):  # 0-9
            hotkey = f'ctrl+shift+{i}'
            keyboard.add_hotkey(hotkey, lambda n=str(i): self.post_key_press_event(n))
            self.hotkey_ids.append(hotkey)

    def post_key_press_event(self, key):
        # Post a custom key press event to the main thread
        event = KeyPressEvent(key)
        QCoreApplication.postEvent(self, event)

    def customEvent(self, event):
        if isinstance(event, KeyPressEvent):
            self.process_key_press(event.key)

    def process_key_press(self, key):
        for data in self.button_data:
            if data['name'].startswith(f"({key})"):
                if data['name'].find("N/A") != -1:
                    # self.show_alert()
                    self.show_error_message("Error", "No Data Found !!!")
                    break
                else:
                    self.on_button_click(data['row'])
                    break

    def disable_hotkeys(self):
        # Disable all registered hotkeys
        for hotkey in self.hotkey_ids:
            keyboard.remove_hotkey(hotkey)

    def enable_hotkeys(self):
        # Re-register all hotkeys
        self.setup_hotkeys()

    def show_error_message(self, title, message):
        """ Show an error message box """
        msg_box = QMessageBox()
        msg_box.setIcon(QMessageBox.Icon.Critical)
        html_message = "<b style='color: red; font-size:18px;'>" + message + "</b>"
        msg_box.setText(html_message)
        msg_box.setWindowTitle(title)
        msg_box.exec()

if __name__ == '__main__':
    app = QApplication([])
    widget = MyWidget()
    widget.show()
    app.exec()
